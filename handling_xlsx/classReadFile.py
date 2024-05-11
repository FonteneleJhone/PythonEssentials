"""
This class is in charge of handle information from .xlsx files

Author: Jhone Fontenele

Year: 2024
"""
__author__="Jhone Fontenele"
__year__="2024"

import openpyxl
import time

class ReadFile:
	#
	def __init__(self, company, products):
		self.company = company
		self.products = products
		#
	def find(self):
		get_file = openpyxl.load_workbook('file.xlsx')
		print(f"\n >> Sheets Available: {get_file.sheetnames}")
		#
		productList = get_file["Product"]
		print(f"\n >> The file has {productList.max_row -1} rows available inside the {productList}")
		#
		products_per_supplier = {}
		supplier = {}
		#
		for i in range(1, productList.max_row + 1):
			#print(f"\n >> Scanning the file {i}")
			companyName = productList.cell(i,3).value
			supplier[i] = companyName
			productPrice = productList.cell(i,2).value
			productType = productList.cell(i,1).value
			#
			if companyName in products_per_supplier:
				numb_products = products_per_supplier[companyName]
				products_per_supplier[companyName] = numb_products + 1
			else:
				products_per_supplier[companyName] = 1
		print(f"\n >> Reading data available - Company:{companyName} Price:R${productPrice} Type:{productType}")
		print(f"\n >> The number of products per supplier is {products_per_supplier[companyName]}")
		#
		print(products_per_supplier)
		#
		for j in range(1, productList.max_row +1):
			if self.company == supplier[j]:
				print(f"\n >>The company searched is:{supplier[j]}")
				return
				


